import csv
import io
import re
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .response_cache import get_cached_value, set_cached_value
from ..supabase_client import get_supabase_client
from .schedule_settings import calculate_attendance_snapshot
from .time_utils import is_leave_code, to_minutes


def format_duration(minutes: int | None) -> str:
  total_minutes = max(int(minutes or 0), 0)
  hours = total_minutes // 60
  remaining_minutes = total_minutes % 60
  return f"{hours}:{remaining_minutes:02d}"


def _calculate_total_hours(record: dict | None) -> str:
  if not record:
    return ""

  leave_type = (record.get("leave_type") or "").strip().upper() or None
  time_in_token = (record.get("time_in") or "").strip().upper()
  time_out_token = (record.get("time_out") or "").strip().upper()
  is_ob_record = leave_type == "OB" or time_in_token == "OB" or time_out_token == "OB"
  if leave_type and leave_type != "OB":
    return ""

  schedule_type = str(record.get("schedule_type") or "A").strip().upper()
  required_minutes = 600 if schedule_type == "B" else 480
  category = str(record.get("category") or "").strip().lower()
  record_floor_minutes = 8 * 60 if category == "jo" else 7 * 60

  if is_ob_record:
    if not time_out_token:
      return format_duration(required_minutes)

    effective_time_in_minutes = record_floor_minutes if time_in_token in {"", "OB"} else to_minutes(record.get("time_in"))
    if effective_time_in_minutes is None:
      effective_time_in_minutes = record_floor_minutes
    else:
      effective_time_in_minutes = max(effective_time_in_minutes, record_floor_minutes)

    schedule_end_minutes = 1140 if schedule_type == "B" else 1020
    time_out_minutes = schedule_end_minutes if time_out_token == "OB" else to_minutes(record.get("time_out"))
    if time_out_minutes is None:
      return format_duration(required_minutes)

    gross_minutes = max(time_out_minutes - effective_time_in_minutes, 0)
    worked_minutes = max(gross_minutes - 60, 0)
    total_minutes = max(min(worked_minutes, required_minutes), 0)
    return format_duration(total_minutes)

  time_in_minutes = to_minutes(record.get("time_in"))
  time_out_minutes = to_minutes(record.get("time_out"))
  if time_in_minutes is None or time_out_minutes is None:
    return ""

  time_in_minutes = max(time_in_minutes, record_floor_minutes)

  gross_minutes = max(time_out_minutes - time_in_minutes, 0)
  worked_minutes = max(gross_minutes - 60, 0)
  late_minutes = int(record.get("late_minutes") or 0)
  total_minutes = max(min(worked_minutes, required_minutes) - late_minutes, 0)

  return format_duration(total_minutes)


def get_month_range(month: str) -> tuple[str, str]:
  dt = datetime.strptime(month, "%Y-%m")
  start = dt.strftime("%Y-%m-01")

  if dt.month == 12:
    end_dt = dt.replace(year=dt.year + 1, month=1)
  else:
    end_dt = dt.replace(month=dt.month + 1)

  end = end_dt.strftime("%Y-%m-01")
  return start, end


def get_enriched_attendance(month: str | None = None) -> list[dict]:
  cache_key = f"reports:enriched-attendance:v2:{(month or 'all').strip() or 'all'}"
  cached_rows = get_cached_value(cache_key)
  if cached_rows is not None:
    return cached_rows

  supabase = get_supabase_client()
  attendance_query = supabase.table("attendance").select("*")

  if month:
    start, end = get_month_range(month)
    attendance_query = attendance_query.gte("date", start).lt("date", end)

  attendance_rows = attendance_query.execute().data or []
  employee_rows = supabase.table("employees").select("id,name,category").execute().data or []
  employee_map = {row["id"]: row for row in employee_rows}

  enriched = []
  for row in attendance_rows:
    employee = employee_map.get(row["employee_id"], {})
    snapshot = {}
    date_value = str(row.get("date") or "")
    if date_value:
      try:
        snapshot = calculate_attendance_snapshot(
          date_value,
          employee.get("category"),
          row.get("schedule_type"),
          row.get("time_in"),
          row.get("time_out"),
          row.get("leave_type")
        )
      except Exception:
        snapshot = {}

    merged = {
      **row,
      **snapshot,
      "employee_name": employee.get("name", "Unknown Employee"),
      "category": employee.get("category", "unknown")
    }
    enriched.append(merged)

  rows = sorted(enriched, key=lambda item: (item.get("date", ""), item.get("employee_name", "")), reverse=True)
  set_cached_value(cache_key, rows, ttl_seconds=300.0)
  return rows


def build_monthly_summary(rows: list[dict]) -> list[dict]:
  summary: dict[tuple[str, str], dict] = {}

  for row in rows:
    key = (row["employee_name"], row["category"])
    if key not in summary:
      summary[key] = {
        "employee_name": row["employee_name"],
        "category": row["category"],
        "days_worked": 0,
        "total_late_minutes": 0,
        "total_undertime_minutes": 0,
        "total_overtime_minutes": 0
      }

    summary[key]["days_worked"] += 1
    summary[key]["total_late_minutes"] += int(row.get("late_minutes") or 0)
    summary[key]["total_undertime_minutes"] += int(row.get("undertime_minutes") or 0)
    summary[key]["total_overtime_minutes"] += int(row.get("overtime_minutes") or 0)

  return sorted(summary.values(), key=lambda item: item["employee_name"])


def build_late_report(rows: list[dict]) -> list[dict]:
  return [
    {
      "date": row.get("date"),
      "employee_name": row.get("employee_name"),
      "category": row.get("category"),
      "late_minutes": row.get("late_minutes", 0)
    }
    for row in rows
    if int(row.get("late_minutes") or 0) > 0
  ]


def build_overtime_report(rows: list[dict]) -> list[dict]:
  return [
    {
      "date": row.get("date"),
      "employee_name": row.get("employee_name"),
      "category": row.get("category"),
      "overtime_minutes": row.get("overtime_minutes", 0)
    }
    for row in rows
    if int(row.get("overtime_minutes") or 0) > 0
  ]


def export_csv(rows: list[dict]) -> bytes:
  output = io.StringIO()
  writer = csv.writer(output)
  writer.writerow([
    "Date",
    "Employee",
    "Category",
    "Time In",
    "Time Out",
    "Late (h:mm)",
    "Undertime (h:mm)",
    "Leave",
    "Schedule"
  ])

  for row in rows:
    writer.writerow([
      row.get("date"),
      row.get("employee_name"),
      row.get("category"),
      row.get("time_in"),
      row.get("time_out"),
      format_duration(row.get("late_minutes")),
      format_duration(row.get("undertime_minutes")),
      row.get("leave_type"),
      row.get("schedule_type")
    ])

  return output.getvalue().encode("utf-8")


def export_xlsx(rows: list[dict]) -> bytes:
  workbook = Workbook()
  sheet = workbook.active
  sheet.title = "Attendance Report"

  sheet.append([
    "Date",
    "Employee",
    "Category",
    "Time In",
    "Time Out",
    "Late (h:mm)",
    "Undertime (h:mm)",
    "Leave",
    "Schedule"
  ])

  for row in rows:
    sheet.append([
      row.get("date"),
      row.get("employee_name"),
      row.get("category"),
      row.get("time_in"),
      row.get("time_out"),
      format_duration(row.get("late_minutes")),
      format_duration(row.get("undertime_minutes")),
      row.get("leave_type"),
      row.get("schedule_type")
    ])

  binary = io.BytesIO()
  workbook.save(binary)
  return binary.getvalue()


def _display_master_sheet_value(row: dict, field: str) -> str:
  leave_type = (row.get("leave_type") or "").strip().upper() or None
  value = row.get(field)

  if field == "time_in":
    if leave_type:
      return leave_type
    if is_leave_code(value):
      return value

  if field == "time_out" and is_leave_code(value):
    return value

  return value or ""


def _display_detail_sheet_value(row: dict, field: str) -> str:
  leave_type = (row.get("leave_type") or "").strip().upper() or None
  value = row.get(field)

  if leave_type:
    if field in {"time_in", "time_out"}:
      return ""
    if field == "leave_type":
      return leave_type

  if field in {"time_in", "time_out"} and is_leave_code(value):
    return ""

  if field == "leave_type":
    return (value or "").strip().upper()

  return value or ""


def _compose_employee_display_name(employee: dict) -> str:
  parts = [
    employee.get("first_name"),
    employee.get("second_name"),
    employee.get("last_name"),
    employee.get("extension")
  ]
  full_name = " ".join(part.strip() for part in parts if isinstance(part, str) and part.strip())
  if full_name:
    return " ".join(full_name.split())

  return (employee.get("display_name") or employee.get("name") or "Unknown Employee").strip() or "Unknown Employee"


def _format_date_label(value: str) -> str:
  parsed = datetime.strptime(value, "%Y-%m-%d")
  return f"{parsed.strftime('%B')} {parsed.day}, {parsed.year}"


def _format_period_label(date_from: str, date_to: str) -> str:
  if not date_from or not date_to:
    return "Selected Period"

  if date_from == date_to:
    return _format_date_label(date_from)

  return f"{_format_date_label(date_from)} - {_format_date_label(date_to)}"


def _normalize_sheet_name(value: str | None) -> str:
  cleaned = re.sub(r"[\[\]\:\*\?/\\]", " ", (value or "")).strip()
  cleaned = re.sub(r"\s+", " ", cleaned)
  return cleaned or "Sheet"


def _build_unique_sheet_name(value: str | None, existing_names: set[str]) -> str:
  base_name = _normalize_sheet_name(value).upper()
  candidate = base_name[:31]
  suffix = 2

  while candidate in existing_names:
    suffix_text = f"_{suffix}"
    base_length = max(1, 31 - len(suffix_text))
    candidate = f"{base_name[:base_length].rstrip()}{suffix_text}"
    suffix += 1

  existing_names.add(candidate)
  return candidate


def _format_employee_date(value: str) -> str:
  parsed = datetime.strptime(value, "%Y-%m-%d")
  return parsed.strftime("%m/%d/%y")


def _format_employee_time(value: str | None) -> str:
  token = (value or "").strip()
  if not token:
    return ""

  if re.fullmatch(r"\d{2}:\d{2}", token):
    hour, minute = token.split(":")
    return f"{int(hour)}:{minute}"

  return token


def _display_placeholder(value: str | None, fallback: str = "N/A") -> str:
  token = (value or "").strip()
  return token or fallback


def _compose_employee_first_name(employee: dict, fallback: str = "N/A") -> str:
  first_name = (employee.get("first_name") or "").strip()
  second_name = (employee.get("second_name") or "").strip()

  if not first_name:
    return _display_placeholder(employee.get("display_name") or employee.get("name"), fallback)

  if not second_name:
    return first_name

  if len(second_name) <= 2:
    return f"{first_name} {second_name.rstrip('.').upper()}.".strip()

  if len(second_name.split()) == 1:
    return f"{first_name} {second_name[0].upper()}.".strip()

  return f"{first_name} {second_name}".strip()


def _compose_employee_signature_name(employee: dict, fallback: str = "N/A") -> str:
  first_name = (employee.get("first_name") or "").strip()
  last_name = (employee.get("last_name") or employee.get("surname") or "").strip()

  if first_name and last_name:
    return f"{first_name} {last_name}".strip()

  if first_name:
    return first_name

  if last_name:
    return last_name

  return _display_placeholder(employee.get("display_name") or employee.get("name"), fallback)


def _write_employee_page_header(sheet, employee: dict, period_label: str) -> tuple[int, int]:
  end_column = 8
  last_column = get_column_letter(end_column)
  title_fill = PatternFill("solid", fgColor="FFFFFF")
  title_font = Font(color="1F4E79", bold=True, size=11)
  sub_font = Font(color="1F4E79", bold=True, size=10)
  main_font = Font(color="444444", bold=True, size=15)
  date_font = Font(color="666666", bold=True, size=10)
  label_font = Font(color="222222", bold=False, size=10)
  value_font = Font(color="111111", bold=True, size=10)

  sheet.sheet_view.showGridLines = False
  sheet.page_setup.orientation = "portrait"
  sheet.page_setup.fitToWidth = 1
  sheet.page_setup.fitToHeight = 0
  sheet.sheet_properties.tabColor = "0F6B6B"
  sheet.freeze_panes = "A14"
  sheet.print_title_rows = "$1:$13"

  sheet.merge_cells(f"A2:{last_column}2")
  sheet.merge_cells(f"A3:{last_column}3")
  sheet.merge_cells(f"A5:{last_column}5")
  sheet.merge_cells(f"A6:{last_column}6")
  sheet.cell(2, 1, "TECHNICAL EDUCATION AND SKILLS DEVELOPMENT AUTHORITY (TESDA)")
  sheet.cell(3, 1, "National Capital Region - MuniPalasTaPat")
  sheet.cell(5, 1, "DAILY TIME RECORD")
  sheet.cell(6, 1, period_label)

  for row_number, font, height in [
    (2, title_font, 18),
    (3, sub_font, 16),
    (5, main_font, 22),
    (6, date_font, 16)
  ]:
    cell = sheet.cell(row_number, 1)
    cell.font = font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[row_number].height = height

  sheet.merge_cells("A8:A8")
  sheet.cell(8, 1, "Employee No.:")
  sheet.cell(8, 1).font = label_font
  sheet.cell(8, 1).alignment = Alignment(horizontal="left", vertical="center")

  sheet.merge_cells(f"B8:{last_column}8")
  sheet.cell(8, 2, _display_placeholder(str(employee.get("employee_no") or employee.get("id") or "").strip()))
  sheet.cell(8, 2).font = value_font
  sheet.cell(8, 2).alignment = Alignment(horizontal="left", vertical="center")

  sheet.cell(9, 1, "Last Name:")
  sheet.cell(9, 1).font = label_font
  sheet.cell(9, 1).alignment = Alignment(horizontal="left", vertical="center")
  sheet.merge_cells(f"B9:{last_column}9")
  sheet.cell(9, 2, _display_placeholder((employee.get("last_name") or employee.get("surname") or employee.get("name") or "").strip().upper()))
  sheet.cell(9, 2).font = value_font
  sheet.cell(9, 2).alignment = Alignment(horizontal="left", vertical="center")

  sheet.cell(10, 1, "First Name:")
  sheet.cell(10, 1).font = label_font
  sheet.cell(10, 1).alignment = Alignment(horizontal="left", vertical="center")
  sheet.merge_cells(f"B10:{last_column}10")
  sheet.cell(10, 2, _compose_employee_first_name(employee, "N/A").upper())
  sheet.cell(10, 2).font = value_font
  sheet.cell(10, 2).alignment = Alignment(horizontal="left", vertical="center")

  sheet.cell(11, 1, "Office:")
  sheet.cell(11, 1).font = label_font
  sheet.cell(11, 1).alignment = Alignment(horizontal="left", vertical="center")
  sheet.merge_cells(f"B11:{last_column}11")
  office_value = _display_placeholder((employee.get("office") or "").strip().upper())
  sheet.cell(11, 2, office_value)
  sheet.cell(11, 2).font = value_font
  sheet.cell(11, 2).alignment = Alignment(horizontal="left", vertical="center")

  for row_number in range(8, 12):
    sheet.row_dimensions[row_number].height = 18

  return 13, end_column


def _write_certification_footer(sheet, start_row: int, end_column: int, footer_data: dict | None = None) -> None:
  footer_data = footer_data or {}
  left_name = (footer_data.get("left_name") or "ATHENA B. VICENTE").strip()
  right_name = (footer_data.get("right_name") or "GERARDO A. MERCADO").strip()
  right_title = (footer_data.get("right_title") or "Head of Office").strip()

  statement = (
    footer_data.get("statement")
    or "I CERTIFY on my honor that the above is a true and correct report of the hours of work performed, record of which was made daily at the time of arrival at and departure from office."
  ).strip()

  left_end = max(3, min(end_column - 2, (end_column // 2) - 1))
  right_start = max(left_end + 2, end_column - max(2, end_column // 3) + 1)
  if right_start > end_column:
    right_start = min(end_column, left_end + 2)

  border_side = Side(style="thin", color="B7B7B7")
  top_border = Border(top=border_side)

  sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_column)
  statement_cell = sheet.cell(start_row, 1, statement)
  statement_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  statement_cell.font = Font(size=10)
  sheet.row_dimensions[start_row].height = 34

  sheet.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=left_end)
  sheet.merge_cells(start_row=start_row + 2, start_column=right_start, end_row=start_row + 2, end_column=end_column)
  left_name_cell = sheet.cell(start_row + 2, 1, left_name)
  right_name_cell = sheet.cell(start_row + 2, right_start, right_name)
  left_name_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  right_name_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  left_name_cell.font = Font(bold=True, size=11)
  right_name_cell.font = Font(bold=True, size=11)
  sheet.row_dimensions[start_row + 2].height = 22

  sheet.merge_cells(start_row=start_row + 3, start_column=1, end_row=start_row + 3, end_column=left_end)
  sheet.merge_cells(start_row=start_row + 3, start_column=right_start, end_row=start_row + 3, end_column=end_column)
  left_label_cell = sheet.cell(start_row + 3, 1, "Name/Signature")
  right_label_cell = sheet.cell(start_row + 3, right_start, f"{right_title}\nName/Signature")
  left_label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  right_label_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  left_label_cell.font = Font(size=10)
  right_label_cell.font = Font(size=10)
  sheet.row_dimensions[start_row + 3].height = 26

  for column_index in range(1, end_column + 1):
    for row_index in range(start_row, start_row + 4):
      sheet.cell(row_index, column_index).border = top_border

  sheet.print_area = f"A1:{get_column_letter(end_column)}{start_row + 3}"


def _write_master_sheet(sheet, sheet_data: dict) -> None:
  sheet.sheet_view.showGridLines = False
  sheet.page_setup.orientation = "landscape"
  sheet.page_setup.fitToWidth = 1
  sheet.page_setup.fitToHeight = 0
  sheet.sheet_properties.tabColor = "0F6B6B"

  employees = sheet_data.get("employees") or []
  dates = sheet_data.get("dates") or []
  records = sheet_data.get("records") or []
  title = sheet_data.get("title") or "Master Record Sheet"

  record_map = {
    (row.get("date"), row.get("employee_id")): row
    for row in records
    if row.get("date") is not None and row.get("employee_id") is not None
  }

  total_columns = max(2 + len(employees) * 2, 2)
  sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
  sheet.cell(1, 1, title)

  dark_fill = PatternFill("solid", fgColor="0F6B6B")
  header_fill = PatternFill("solid", fgColor="DCEFD9")
  subheader_fill = PatternFill("solid", fgColor="ECF6E7")
  date_fill = PatternFill("solid", fgColor="F6E3D2")
  day_fill = PatternFill("solid", fgColor="EFEFEF")
  weekend_fill = PatternFill("solid", fgColor="D7D7D7")
  monday_fill = PatternFill("solid", fgColor="F5C148")
  white_fill = PatternFill("solid", fgColor="FFFFFF")
  thin_side = Side(style="thin", color="7F8C91")
  thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

  title_cell = sheet.cell(1, 1)
  title_cell.fill = dark_fill
  title_cell.font = Font(color="FFFFFF", bold=True, size=14)
  title_cell.alignment = Alignment(horizontal="left", vertical="center")
  sheet.row_dimensions[1].height = 24
  sheet.row_dimensions[2].height = 8

  sheet.merge_cells(start_row=3, start_column=1, end_row=4, end_column=1)
  sheet.merge_cells(start_row=3, start_column=2, end_row=4, end_column=2)
  sheet.cell(3, 1, "DATE")
  sheet.cell(3, 2, "DAY")

  sheet.cell(3, 1).fill = date_fill
  sheet.cell(3, 2).fill = day_fill
  sheet.cell(3, 1).font = Font(bold=True)
  sheet.cell(3, 2).font = Font(bold=True)
  sheet.cell(3, 1).alignment = Alignment(horizontal="center", vertical="center")
  sheet.cell(3, 2).alignment = Alignment(horizontal="center", vertical="center")
  sheet.cell(4, 1).fill = date_fill
  sheet.cell(4, 2).fill = day_fill

  for index, employee in enumerate(employees):
    start_column = 3 + (index * 2)
    surname = (employee.get("surname") or employee.get("last_name") or employee.get("name") or "").strip().upper()

    sheet.merge_cells(start_row=3, start_column=start_column, end_row=3, end_column=start_column + 1)
    sheet.cell(3, start_column, surname)
    sheet.cell(3, start_column).fill = header_fill
    sheet.cell(3, start_column).font = Font(bold=True)
    sheet.cell(3, start_column).alignment = Alignment(horizontal="center", vertical="center")

    sheet.cell(4, start_column, "TIME IN")
    sheet.cell(4, start_column + 1, "TIME OUT")
    sheet.cell(4, start_column).fill = subheader_fill
    sheet.cell(4, start_column + 1).fill = subheader_fill
    sheet.cell(4, start_column).font = Font(bold=True)
    sheet.cell(4, start_column + 1).font = Font(bold=True)
    sheet.cell(4, start_column).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(4, start_column + 1).alignment = Alignment(horizontal="center", vertical="center")

    sheet.column_dimensions[get_column_letter(start_column)].width = 11
    sheet.column_dimensions[get_column_letter(start_column + 1)].width = 11

  sheet.column_dimensions["A"].width = 12
  sheet.column_dimensions["B"].width = 14
  sheet.freeze_panes = "C5"
  sheet.print_title_rows = "$1:$4"

  for row_index, date_info in enumerate(dates, start=5):
    row_date = date_info.get("date")
    weekday = date_info.get("weekday") or ""
    is_weekend = bool(date_info.get("is_weekend"))
    is_monday = bool(date_info.get("is_monday"))
    row_fill = weekend_fill if is_weekend else monday_fill if is_monday else white_fill

    sheet.cell(row_index, 1, date_info.get("label") or row_date)
    sheet.cell(row_index, 2, weekday)
    sheet.cell(row_index, 1).fill = row_fill
    sheet.cell(row_index, 2).fill = row_fill
    sheet.cell(row_index, 1).font = Font(bold=True if is_weekend or is_monday else False)
    sheet.cell(row_index, 2).font = Font(bold=True if is_weekend or is_monday else False)
    sheet.cell(row_index, 1).alignment = Alignment(horizontal="center", vertical="center")
    sheet.cell(row_index, 2).alignment = Alignment(horizontal="left", vertical="center")

    for index, employee in enumerate(employees):
      start_column = 3 + (index * 2)
      record = record_map.get((row_date, employee.get("id")), {})

      time_in_cell = sheet.cell(row_index, start_column, _display_master_sheet_value(record, "time_in"))
      time_out_cell = sheet.cell(row_index, start_column + 1, _display_master_sheet_value(record, "time_out"))

      time_in_cell.fill = white_fill
      time_out_cell.fill = white_fill
      time_in_cell.alignment = Alignment(horizontal="center", vertical="center")
      time_out_cell.alignment = Alignment(horizontal="center", vertical="center")

  for row in sheet.iter_rows(min_row=1, max_row=4 + len(dates), min_col=1, max_col=total_columns):
    for cell in row:
      cell.border = thin_border


def _write_employee_sheet(sheet, employee: dict, dates: list[dict], records: list[dict], period_label: str, footer_data: dict | None = None) -> None:
  employee_id = employee.get("id")
  employee_records = [row for row in records if row.get("employee_id") == employee_id]
  record_map = {
    row.get("date"): row
    for row in employee_records
    if row.get("date") is not None
  }

  detail_rows: list[dict] = []
  for date_info in dates:
    row_date = date_info.get("date")
    record = record_map.get(row_date)
    time_in_value = _format_employee_time(_display_master_sheet_value(record or {}, "time_in"))
    time_out_value = _format_employee_time(_display_master_sheet_value(record or {}, "time_out"))
    late_value = format_duration(record.get("late_minutes")) if record else ""
    undertime_value = format_duration(record.get("undertime_minutes")) if record else ""
    total_hours_value = _calculate_total_hours(record)

    detail_rows.append(
      {
        "date": _format_employee_date(row_date),
        "day": date_info.get("weekday") or "",
        "time_in": time_in_value,
        "time_out": time_out_value,
        "late": late_value,
        "undertime": undertime_value,
        "remarks": (record.get("remarks") or "").strip() if record else "",
        "total_hours": total_hours_value,
      }
    )

  start_row, end_column = _write_employee_page_header(sheet, employee, period_label)
  header_fill = PatternFill("solid", fgColor="3E4A5C")
  header_font = Font(color="FFFFFF", bold=True, size=10)
  row_fill_odd = PatternFill("solid", fgColor="F2F4F7")
  row_fill_even = PatternFill("solid", fgColor="FFFFFF")
  thin_side = Side(style="thin", color="B7B7B7")
  thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

  sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
  sheet.cell(start_row, 1, "DATE")
  sheet.cell(start_row, 1).fill = header_fill
  sheet.cell(start_row, 1).font = header_font
  sheet.cell(start_row, 1).alignment = Alignment(horizontal="center", vertical="center")

  for column_index, header_text in enumerate(["TIME-IN", "TIME-\nOUT", "LATE", "UNDERTIME", "REMARKS", "TOTAL\nHOURS"], start=3):
    cell = sheet.cell(start_row, column_index, header_text)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

  sheet.column_dimensions["A"].width = 12
  sheet.column_dimensions["B"].width = 13
  sheet.column_dimensions["C"].width = 10
  sheet.column_dimensions["D"].width = 10
  sheet.column_dimensions["E"].width = 10
  sheet.column_dimensions["F"].width = 10
  sheet.column_dimensions["G"].width = 17
  sheet.column_dimensions["H"].width = 12
  sheet.row_dimensions[start_row].height = 22

  data_start_row = start_row + 1
  for row_offset, row_data in enumerate(detail_rows, start=data_start_row):
    row_fill = row_fill_even if row_offset % 2 == 0 else row_fill_odd
    values = [
      row_data["date"],
      row_data["day"],
      row_data["time_in"],
      row_data["time_out"],
      row_data["late"],
      row_data["undertime"],
      row_data["remarks"],
      row_data["total_hours"]
    ]

    for column_index, value in enumerate(values, start=1):
      cell = sheet.cell(row_offset, column_index, value)
      cell.fill = row_fill
      cell.border = thin_border
      cell.font = Font(size=9)
      cell.alignment = Alignment(
        horizontal="center" if column_index != 7 else "left",
        vertical="center"
      )

    sheet.row_dimensions[row_offset].height = 18

  footer_start_row = data_start_row + len(detail_rows) + 2
  resolved_footer_data = {
    **(footer_data or {}),
    "left_name": (footer_data or {}).get("left_name") or _compose_employee_signature_name(employee, "N/A").upper(),
    "right_name": (footer_data or {}).get("right_name"),
    "right_title": (footer_data or {}).get("right_title"),
    "statement": (footer_data or {}).get("statement")
  }
  _write_certification_footer(
    sheet,
    footer_start_row,
    end_column,
    resolved_footer_data
  )


def export_master_sheet_xlsx(sheet_data: dict) -> bytes:
  workbook = Workbook()
  sheet = workbook.active
  sheet.title = "Master Sheet"

  _write_master_sheet(sheet, sheet_data)

  employees = sheet_data.get("employees") or []
  dates = sheet_data.get("dates") or []
  records = sheet_data.get("records") or []
  period_label = _format_period_label(sheet_data.get("date_from") or "", sheet_data.get("date_to") or "")

  existing_names = {sheet.title}
  for employee in employees:
    base_name = employee.get("surname") or employee.get("last_name") or employee.get("name")
    sheet_name = _build_unique_sheet_name(base_name, existing_names)
    detail_sheet = workbook.create_sheet(title=sheet_name)
    _write_employee_sheet(detail_sheet, employee, dates, records, period_label, sheet_data.get("footer"))

  binary = io.BytesIO()
  workbook.save(binary)
  return binary.getvalue()
